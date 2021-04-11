# ------- 3rd party imports -------
from flask import Blueprint, request, Response
from flask import Blueprint, request, jsonify
import openpyxl
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from app.server.db.helper import helper
from app.server.db.models import *
from app.server.utils.time import TimeUtils

XLSX_DIR = '/in/'

# ------- local imports -------
rates_blueprint = Blueprint('rates_blueprint', __name__)


@rates_blueprint.route("/rates", methods=['GET'])
def download_rates_to_xml():
    output_dir = os.path.join('app', 'downloads')
    output_file = 'rates.xlsx'
    data = helper.get_all(Rate)
    wb = Workbook()
    ws = wb.active
    if not os.path.isdir(output_dir):
        os.mkdir(output_dir)
    headers = ['Product', 'Rate', 'Scope']
    for col_num, header in enumerate(headers):
        ws.cell(row=1, column=col_num + 1).value = header
    for row_num, row in enumerate(data):
        real_row_num = row_num + 2
        ws.cell(row=real_row_num, column=1).value = row.product_name
        ws.cell(row=real_row_num, column=2).value = row.rate
        ws.cell(row=real_row_num, column=3).value = row.scope

    return Response(save_virtual_workbook(wb), headers={'Content-Disposition': 'attachment; filename=rates.xlsx',
                                                        'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                                        })

    xlsx_file = Path(os.getcwd() + '/app/static/rates.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    title = True
    for row in sheet.iter_rows():
        if title:  # first row is title we skip it
            title = False
            continue

        record = helper.get_one(Rate, product_name=row[0].value)
        if record is not None:
            print(f'value {row[0].value} already exists, updating...')
            record.rate = row[1].value
            record.scope = row[2].value
            helper.commit_changes()
        else:
            helper.add_instance(Rate,
                                product_name=row[0].value,
                                rate=row[1].value,
                                scope=row[2].value)

    print("hello from POST/rates")
    return 'OK', 200


@rates_blueprint.route("/rates", methods=['POST'])
def upload_xml_data():
    data = (request.form and request.form["file"]) or (request.json and request.json.get("file"))

    if data is None or not os.path.isfile('/in/' + data):
        return 'Bad parameters, expected {"file":"<name>"} or "file=<name>"\n', 400
    xlsx_file = Path('/in/' + data)
    try:
        wb_obj = openpyxl.load_workbook(xlsx_file)
    except:
        return f'Bad file {data}, unreadable or wrong format\n', 500
    sheet = wb_obj.active

    title = True
    for row in sheet.iter_rows():
        if title:  # first row is title we skip it
            title = False
            continue

        record = helper.get_one(Rate, product_name=row[0].value)
        if record is not None:
            print(f'value {row[0].value} already exists, updating...')
            record.rate = row[1].value
            record.scope = row[2].value
            helper.commit_changes()
        else:
            helper.add_instance(Rate, product_name=row[0].value,
                                rate=row[1].value,
                                scope=row[2].value)

        TimeUtils.set_file_last_modified(XLSX_DIR + data, TimeUtils.get_now())

    return 'OK\n', 200
